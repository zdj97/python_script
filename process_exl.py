#!/user/bin/env python
# coding=utf-8
'''
@project : 语音分离
@author  : zmj1997
#@file   : process_exl.py
#@ide    : PyCharm
#@time   : 2022-10-19 17:06:36
'''

import openpyxl
from openpyxl.styles import Font  # 导入字体模块
from openpyxl.styles import PatternFill  # 导入填充模块
from openpyxl import load_workbook
# def Textcolor (file_name, title):
#     wk = openpyxl.load_workbook (file_name)  # 加载已经存在的excel
#     sheet1 = wk[title]  # wk[wk_name[0]]#title名称
#     for i in range (10):
#         # Color=['c6efce','006100']#绿
#         # Color = ['ffc7ce', '9c0006']  #红
#         # Color = ['ffeb9c', '9c6500']  # 黄
#         Color = ['ffffff', '000000']  # 黑白
#         fille = PatternFill ('solid', fgColor = Color[0])  # 设置填充颜色为 橙色
#         font = Font (u'微软雅黑', size = 11, bold = True, italic = False, strike = False, color = Color[1])  # 设置字体样式
#         sheet1.cell (row = i + 2, column = 8, value = "").fill = fille  # 序列
#         sheet1.cell (row = i + 2, column = 8, value = "哈哈").font = font  # 序列
#     wk.save (file_name)  # 保存excel
#
#
# file_name, title = 's.xlsx', 'Sheet1'
#
# Textcolor (file_name, title)


colors = ['FFB6C1', '800080', '6A5ACD', '8FBC8F', 'FFE4C4', 'D3D3D3']
file_name, title = './中保转写结果.xlsx', 'Sheet1'
wk = openpyxl.load_workbook (file_name)  # 加载已经存在的excel
sheet1 = wk[title]  # wk[wk_name[0]]#title名称
R = sheet1.max_row
print(R)
temp = []
start_index = end_index = 0
for i in range (2, R + 1):
    a = i % 6
    # print(a)
    if not temp:
        temp.append (sheet1.cell (row = i, column = 1).value)
        start_index = i
        # print(start_index)
    else:
        if sheet1.cell (row = i, column = 1).value in temp:
            continue
        else:
            end_index = i - 1
            temp = []
    if end_index > start_index:
        if start_index == 2:
            for x in range (start_index, end_index + 1):
                for y in range(1,7):
                    fille = PatternFill ('solid', fgColor = colors[a])
                    sheet1.cell (row = x, column =y).fill = fille
                # sheet1.cell (row = x, column = 6).fill = colors[a]
                # pass
        else:
            print (start_index, end_index)
            for x in range (start_index-1, end_index + 1):
                for y in range(1,7):
                    fille = PatternFill ('solid', fgColor = colors[a])
                    sheet1.cell (row = x, column =y).fill = fille
                # sheet1.cell (row = x, column = 6).fill = colors[a]
                # pass
wk.save ('./ttt.xlsx')




# fille = PatternFill ('solid', fgColor = colors[a])
# sheet1.cell(row=i,column=2).fill=colors[a]
# wk.save(file_name)
